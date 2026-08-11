#!/usr/bin/env python3
"""Build rates.json from the newest selected Wheelsys spreadsheets.

Families supported:
- BK
- FCI
- VAN

Expected columns:
Group, Pickup start, Pickup end, Rate zone, Booking start, Booking end,
1 per day, 2 per day, 3 per day, 4 - 6 per day, 7 per day,
8 - 13 per day, 14 - 29 per day, 30+ per day.

Locations are inferred from Rate zone: LXA -> Lisboa, OPT -> Porto, FAO -> Faro.
"""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SELECTED = ROOT / "data" / "selected"
OUT_DIR = ROOT / "public" / "data"
OUT_JSON = OUT_DIR / "rates.json"
BUILD_MANIFEST = ROOT / "build_manifest.json"

ZONE_TO_LOCATION = {"LXA": "Lisboa", "OPT": "Porto", "FAO": "Faro"}
FAMILIES = ("BK", "FCI", "VAN")

CANONICAL = {
    "group": ["group"],
    "pickupStart": ["pickup start", "pickupstart"],
    "pickupEnd": ["pickup end", "pickupend"],
    "rateZone": ["rate zone", "ratezone"],
    "bookingStart": ["booking start", "bookingstart"],
    "bookingEnd": ["booking end", "bookingend"],
    "p1": ["1 per day", "1  per day", "1/day"],
    "p2": ["2 per day", "2  per day", "2/day"],
    "p3": ["3 per day", "3  per day", "3/day"],
    "p46": ["4 - 6 per day", "4-6 per day", "4 - 6  per day"],
    "p7": ["7 per day", "7  per day", "7/day"],
    "p813": ["8 - 13 per day", "8-13 per day", "8 - 13  per day"],
    "p1429": ["14 - 29 per day", "14-29 per day", "14 - 29  per day"],
    "p30": ["30+ per day", "30+  per day", "30 plus per day"],
}


def norm(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    return re.sub(r"\s+", " ", text)


def format_date(value: Any, booking: bool = False) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S" if booking else "%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d 00:00:00" if booking else "%d/%m/%Y")
    return str(value).strip()


def tariff_from_filename(path: Path) -> str:
    name = path.stem.upper()
    if "COMMERCIAL_VAN" in name or "VAN" in name:
        return "VAN"
    if "FCI" in name:
        return "FCI"
    if "BK" in name:
        return "BK"
    raise ValueError(f"Familia de tarifa nao reconhecida: {path.name}")


def choose_sheet(wb):
    for name in wb.sheetnames:
        if norm(name) == "base rates":
            return wb[name]
    return wb[wb.sheetnames[0]]


def map_headers(values: list[Any]) -> dict[str, int]:
    normalized = [norm(v) for v in values]
    result: dict[str, int] = {}
    for key, aliases in CANONICAL.items():
        alias_set = {norm(a) for a in aliases}
        for idx, value in enumerate(normalized):
            if value in alias_set:
                result[key] = idx
                break
    missing = sorted(set(CANONICAL) - set(result))
    if missing:
        raise ValueError("Colunas em falta: " + ", ".join(missing))
    return result


def numeric(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("€", "").replace(" ", "")
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    elif "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    return float(text)


def parse_workbook(path: Path) -> tuple[str, list[dict[str, Any]], list[str]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = choose_sheet(wb)
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        raise ValueError("Folha vazia")

    mapping = map_headers(header)
    tariff = tariff_from_filename(path)
    records: list[dict[str, Any]] = []
    warnings: list[str] = []

    for excel_row, row in enumerate(rows, start=2):
        if not any(v not in (None, "") for v in row):
            continue
        zone = str(row[mapping["rateZone"]] or "").strip().upper()
        location = ZONE_TO_LOCATION.get(zone)
        if not location:
            warnings.append(f"{path.name}: linha {excel_row}: Rate zone desconhecida '{zone}'")
            continue
        group = str(row[mapping["group"]] or "").strip()
        if not group:
            warnings.append(f"{path.name}: linha {excel_row}: Group vazio")
            continue

        records.append({
            "tariff": tariff,
            "location": location,
            "group": group,
            "pickupStart": format_date(row[mapping["pickupStart"]]),
            "pickupEnd": format_date(row[mapping["pickupEnd"]]),
            "bookingStart": format_date(row[mapping["bookingStart"]], booking=True),
            "bookingEnd": format_date(row[mapping["bookingEnd"]]),
            "p1": numeric(row[mapping["p1"]]),
            "p2": numeric(row[mapping["p2"]]),
            "p3": numeric(row[mapping["p3"]]),
            "p46": numeric(row[mapping["p46"]]),
            "p7": numeric(row[mapping["p7"]]),
            "p813": numeric(row[mapping["p813"]]),
            "p1429": numeric(row[mapping["p1429"]]),
            "p30": numeric(row[mapping["p30"]]),
        })

    wb.close()
    return tariff, records, warnings


def main() -> int:
    SELECTED.mkdir(parents=True, exist_ok=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(p for p in SELECTED.iterdir() if p.suffix.lower() in {".xlsx", ".xlsm"})
    if not files:
        print(f"Nenhum Excel encontrado em {SELECTED}")
        return 1

    rates = {
        family: {loc: [] for loc in ZONE_TO_LOCATION.values()}
        for family in FAMILIES
    }
    source_summary = []
    all_warnings: list[str] = []

    for path in files:
        tariff, records, warnings = parse_workbook(path)
        for rec in records:
            clean = {k: v for k, v in rec.items() if k not in {"tariff", "location"}}
            rates[tariff][rec["location"]].append(clean)
        source_summary.append({"file": path.name, "tariff": tariff, "rows": len(records)})
        all_warnings.extend(warnings)

    for tariff in rates.values():
        for records in tariff.values():
            records.sort(key=lambda r: (r["group"], r["pickupStart"], r["pickupEnd"]))

    OUT_JSON.write_text(json.dumps(rates, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    manifest = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "output": str(OUT_JSON.relative_to(ROOT)),
        "sources": source_summary,
        "families": list(FAMILIES),
        "warnings": all_warnings,
    }
    BUILD_MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"Criado: {OUT_JSON}")
    for item in source_summary:
        print(f"- {item['file']}: {item['tariff']} | {item['rows']} linhas")
    if all_warnings:
        print(f"Avisos: {len(all_warnings)} (ver build_manifest.json)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
